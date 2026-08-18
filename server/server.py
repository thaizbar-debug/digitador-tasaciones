"""
Server local del digitador de tasaciones.

Sirve el HTML en / y expone POST /generar que recibe borrador + fotos + banco
y devuelve un ZIP con:
  - informe-<jobName>-<banco>.xlsx (plantilla oficial del banco, con datos
    del borrador transferidos, fotos insertadas en la hoja FOTO, y todos
    los logos e imágenes originales preservados por openpyxl)
  - revisar.md, checklist manual antes de enviar
  - anexo-fotografico.md, listado de fotos con rotulado sugerido
  - manifest.json, metadata del trabajo

Cómo correr:
  python3 -m venv .venv
  .venv/bin/pip install -r server/requirements.txt
  .venv/bin/python server/server.py

O simplemente doble click a abrir.command en la raíz del proyecto.
"""

import io
import tempfile
import warnings
from collections import defaultdict
from pathlib import Path

from flask import Flask, jsonify, request, send_file, send_from_directory
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

PROYECTO_DIR = Path(__file__).resolve().parent.parent
FORMATOS_DIR = PROYECTO_DIR / "formatos"
WEB_DIR = PROYECTO_DIR / "web"

app = Flask(__name__, static_folder=None)


@app.route("/")
def index():
    return send_from_directory(WEB_DIR, "index.html")


@app.route("/<path:path>")
def static_files(path):
    return send_from_directory(WEB_DIR, path)


def find_plantilla(banco: str) -> Path:
    banco_dir = FORMATOS_DIR / banco
    if not banco_dir.is_dir():
        raise FileNotFoundError(f"No existe la carpeta formatos/{banco}")
    xlsx_files = sorted(banco_dir.glob("*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(f"No hay .xlsx en formatos/{banco}")
    return xlsx_files[0]


def merge_borrador_into_plantilla(borrador_path: Path, plantilla_path: Path):
    """Copia valores del borrador a la plantilla, respetando fórmulas oficiales."""
    wb_plantilla = load_workbook(plantilla_path)
    wb_borrador = load_workbook(borrador_path, data_only=True)

    stats = {
        "plantillaFilename": plantilla_path.name,
        "hojasProcesadas": [],
        "hojasIgnoradas": [],
        "hojasFaltantes": [],
        "celdasEscritas": 0,
    }

    borrador_by_trimmed = {ws.title.strip(): ws for ws in wb_borrador.worksheets}
    plantilla_trimmed = {ws.title.strip() for ws in wb_plantilla.worksheets}

    for trimmed, ws in borrador_by_trimmed.items():
        if trimmed not in plantilla_trimmed:
            stats["hojasIgnoradas"].append(ws.title)

    for p_ws in wb_plantilla.worksheets:
        b_ws = borrador_by_trimmed.get(p_ws.title.strip())
        if not b_ws:
            stats["hojasFaltantes"].append(p_ws.title)
            continue

        # Set of coordinates that are inside a merged range but are NOT the anchor.
        # Writing to those raises AttributeError in openpyxl.
        merged_slaves = set()
        for mr in p_ws.merged_cells.ranges:
            for coord in mr.cells:
                if coord != (mr.min_row, mr.min_col):
                    merged_slaves.add(coord)

        written = 0
        for row in b_ws.iter_rows():
            for b_cell in row:
                if b_cell.value is None:
                    continue
                coord = (b_cell.row, b_cell.column)
                if coord in merged_slaves:
                    continue
                p_cell = p_ws.cell(row=b_cell.row, column=b_cell.column)
                if isinstance(p_cell.value, str) and p_cell.value.startswith("="):
                    continue
                try:
                    p_cell.value = b_cell.value
                    written += 1
                except AttributeError:
                    pass  # celda combinada; ignorar
        stats["hojasProcesadas"].append({"hoja": p_ws.title, "celdas": written})
        stats["celdasEscritas"] += written

    return wb_plantilla, stats


def detectar_slots_fotos(ws):
    """
    Detecta los slots de foto en la hoja FOTO buscando el patrón:
    varias celdas fusionadas chicas alineadas en la misma fila (labels
    numéricos debajo de cada foto). Retorna una lista ordenada de tuplas
    (anchor, width_cols, height_rows) para cada slot.
    """
    if not ws.merged_cells.ranges:
        return []

    por_fila = defaultdict(list)
    for mr in ws.merged_cells.ranges:
        # Labels típicos: una sola fila, 2 a 6 columnas de ancho
        if mr.min_row == mr.max_row and 1 <= (mr.max_col - mr.min_col + 1) <= 6:
            por_fila[mr.min_row].append(mr)

    filas_label = sorted([r for r, v in por_fila.items() if len(v) >= 3])
    if not filas_label:
        return []

    slots = []
    prev_row = 1
    for label_row in filas_label:
        top = prev_row + 1
        height = label_row - top
        if height < 3:
            prev_row = label_row
            continue
        for mr in sorted(por_fila[label_row], key=lambda m: m.min_col):
            col_letter = get_column_letter(mr.min_col)
            width_cols = mr.max_col - mr.min_col + 1
            slots.append({
                "anchor": f"{col_letter}{top}",
                "width_cols": width_cols,
                "height_rows": height,
            })
        prev_row = label_row

    return slots


def insert_fotos(wb, foto_paths):
    """
    Inserta las fotos subidas en los slots detectados de la hoja FOTO
    (ordenados de arriba a abajo, izquierda a derecha). Si no se detectan
    slots o hay más fotos que slots, las sobrantes van al final de la hoja.
    """
    result = {"insertadas": 0, "en_slots": 0, "al_final": 0, "nota": ""}
    foto_sheet_name = next(
        (n for n in wb.sheetnames if "FOTO" in n.upper()), None
    )
    if not foto_sheet_name:
        result["nota"] = "No se encontró hoja FOTO en la plantilla"
        return result

    ws = wb[foto_sheet_name]
    slots = detectar_slots_fotos(ws)

    def preparar_imagen(foto_path, target_w_px=None, target_h_px=None):
        pil = PILImage.open(foto_path)
        # Aproximación: ancho de columna ~50 px, alto de fila ~15 px
        if target_w_px and target_h_px:
            pil.thumbnail((target_w_px, target_h_px))
        else:
            pil.thumbnail((600, 450))
        if pil.mode in ("RGBA", "P"):
            pil = pil.convert("RGB")
        buf = io.BytesIO()
        pil.save(buf, format="JPEG", quality=85)
        buf.seek(0)
        return buf

    fallback_start_row = max(ws.max_row, 5) + 3
    fallback_count = 0

    for idx, foto_path in enumerate(foto_paths):
        try:
            if idx < len(slots):
                slot = slots[idx]
                # Aprox 50 px por columna, 15 px por fila
                w = max(80, slot["width_cols"] * 50)
                h = max(60, slot["height_rows"] * 15)
                buf = preparar_imagen(foto_path, w + 100, h + 100)
                img = XLImage(buf)
                img.width = w
                img.height = h
                ws.add_image(img, slot["anchor"])
                result["en_slots"] += 1
            else:
                buf = preparar_imagen(foto_path)
                img = XLImage(buf)
                img.width = 300
                img.height = 225
                row = fallback_start_row + (fallback_count // 2) * 18
                col = "B" if fallback_count % 2 == 0 else "H"
                ws.add_image(img, f"{col}{row}")
                fallback_count += 1
                result["al_final"] += 1
            result["insertadas"] += 1
        except Exception as exc:
            print(f"[fotos] error insertando {foto_path}: {exc}")

    if result["en_slots"] and result["al_final"]:
        result["nota"] = f"{result['en_slots']} fotos en los slots oficiales, {result['al_final']} al final (sobrantes)."
    elif result["en_slots"]:
        result["nota"] = f"{result['en_slots']} fotos colocadas en los slots oficiales de la hoja FOTO."
    else:
        result["nota"] = f"No se detectaron slots; {result['al_final']} fotos al final de la hoja (fila {fallback_start_row}+)."
    return result


@app.route("/generar", methods=["POST"])
def generar():
    job_name = (request.form.get("jobName") or "").strip()
    banco = (request.form.get("banco") or "").strip()

    if not job_name:
        return jsonify({"error": "Falta el nombre del trabajo."}), 400
    if not banco:
        return jsonify({"error": "Falta elegir el banco."}), 400

    borrador_files = request.files.getlist("borrador")
    if not borrador_files:
        return jsonify({"error": "Falta el borrador del tasador."}), 400

    foto_files = request.files.getlist("fotos")

    try:
        plantilla_path = find_plantilla(banco)
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 400

    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        borrador_path = tmp / "borrador.xlsx"
        borrador_files[0].save(borrador_path)

        foto_paths = []
        for pf in foto_files:
            safe = pf.filename.replace("/", "_")
            fp = tmp / f"foto_{len(foto_paths)}_{safe}"
            pf.save(fp)
            foto_paths.append(fp)

        wb, _stats = merge_borrador_into_plantilla(borrador_path, plantilla_path)
        insert_fotos(wb, foto_paths)

        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)

        return send_file(
            out_buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"informe-{job_name}-{banco}.xlsx",
        )


@app.route("/health")
def health():
    return jsonify({"ok": True, "bancos": [d.name for d in FORMATOS_DIR.iterdir() if d.is_dir()]})


if __name__ == "__main__":
    print(f"Digitador de tasaciones. Proyecto: {PROYECTO_DIR}")
    print("Abre http://localhost:5555/ en tu navegador.")
    app.run(host="127.0.0.1", port=5555, debug=False)
